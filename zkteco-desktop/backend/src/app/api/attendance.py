from app.shared.logger import app_logger

from flask import Blueprint, jsonify, request, send_file
from app.services.device_service import get_zk_service
from app.services.attendance_sync_service import attendance_sync_service
from app.services.scheduler_service import scheduler_service
from app.services.attendance_cleanup_service import attendance_cleanup_service
from app.repositories import attendance_repo, user_repo
from flask import current_app
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

bp = Blueprint("attendance", __name__, url_prefix="/")


def get_service():
    return get_zk_service()


def _map_user_names_to_logs(logs):
    """Helper function to map user names to attendance logs."""
    users = user_repo.get_all()
    user_map = {user.user_id: user.name for user in users}

    logs_with_names = []
    for log in logs:
        log_dict = log.to_dict()
        log_dict["name"] = user_map.get(log.user_id, "Unknown User")
        logs_with_names.append(log_dict)
    return logs_with_names


def _map_user_details_to_logs(logs):
    """Helper function to map user names, avatars, and employee details to attendance logs."""
    users = user_repo.get_all()
    # Use a composite key (user_id, serial_number) to handle duplicate user_ids across devices
    user_map = {
        (user.user_id, user.serial_number): {
            "name": user.name,
            "avatar_url": getattr(user, "avatar_url", None),
            "full_name": getattr(user, "full_name", None),
            "employee_code": getattr(user, "employee_code", None),
            "position": getattr(user, "position", None),
            "department": getattr(user, "department", None),
            "notes": getattr(user, "notes", None),
            "employee_object": getattr(
                user, "employee_object", None
            ),  # Add employee_object
            "gender": getattr(user, "gender", None),  # Add gender
            "hire_date": getattr(user, "hire_date", None),  # Add hire_date
        }
        for user in users
    }

    enriched_logs = []
    for log in logs:
        log_dict = log.to_dict()
        # Use the composite key for lookup
        user_info = user_map.get((log.user_id, log.serial_number))

        if user_info:
            log_dict["name"] = user_info["name"]
            log_dict["avatar_url"] = user_info["avatar_url"]
            log_dict["full_name"] = user_info["full_name"]
            log_dict["employee_code"] = user_info["employee_code"]
            log_dict["position"] = user_info["position"]
            log_dict["department"] = user_info["department"]
            log_dict["notes"] = user_info["notes"]
            log_dict["employee_object"] = user_info[
                "employee_object"
            ]  # Add employee_object
            log_dict["gender"] = user_info["gender"]  # Add gender
            log_dict["hire_date"] = user_info["hire_date"]  # Add hire_date
        else:
            log_dict["name"] = "Unknown User"
            log_dict["avatar_url"] = None
            log_dict["full_name"] = None
            log_dict["employee_code"] = None
            log_dict["position"] = None
            log_dict["department"] = None
            log_dict["notes"] = None
            log_dict["employee_object"] = None  # Add employee_object
            log_dict["gender"] = None  # Add gender
            log_dict["hire_date"] = None  # Add hire_date

        enriched_logs.append(log_dict)

    return enriched_logs


@bp.route("/attendance", methods=["GET"])
def get_attendance():
    """Get stored attendance logs with pagination and filtering"""
    try:
        # Query parameters
        device_id = request.args.get("device_id")
        limit = int(request.args.get("limit", 100))
        offset = int(request.args.get("offset", 0))
        user_id = request.args.get("user_id")
        date_str = request.args.get("date")  # Format: YYYY-MM-DD

        # Validate limit
        limit = min(limit, 1000)  # Max 1000 records per request

        # Parse date if provided
        start_date = None
        end_date = None
        if date_str:
            try:
                target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                start_date = datetime.combine(target_date, datetime.min.time())
                end_date = datetime.combine(target_date, datetime.max.time())
            except ValueError:
                return jsonify(
                    {"success": False, "error": "Invalid date format. Use YYYY-MM-DD"}
                ), 400

        if user_id:
            logs = attendance_repo.get_by_user(user_id, limit)
        else:
            logs = attendance_repo.get_all(
                device_id, limit, offset, start_date, end_date
            )

        data = _map_user_details_to_logs(logs)

        return jsonify(
            {
                "success": True,
                "data": data,
                "pagination": {"limit": limit, "offset": offset, "count": len(logs)},
            }
        )

    except Exception as e:
        app_logger.error(f"Error getting attendance logs: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/attendance/sync", methods=["POST"])
def sync_attendance():
    """Sync attendance logs from device to database"""
    try:
        app_logger.info("Syncing attendance from device")
        result = get_service().get_attendance()

        if isinstance(result, dict) and "records" in result:
            attendances = result["records"]
            sync_stats = result.get("sync_stats", {})
        else:
            attendances = result
            sync_stats = None

        if not attendances:
            return jsonify(
                {
                    "message": "Không có dữ liệu chấm công mới trên thiết bị",
                    "data": [],
                    "sync_stats": sync_stats,
                }
            )

        # The get_attendance service already saves to DB, so we just return the stats
        response = {
            "message": "Đồng bộ dữ liệu chấm công thành công",
            "sync_stats": sync_stats,
        }

        return jsonify(response)
    except Exception as e:
        error_message = f"Lỗi khi đồng bộ dữ liệu chấm công: {str(e)}"
        current_app.logger.error(error_message)
        return jsonify({"message": error_message}), 500


# Sync Management APIs


@bp.route("/attendance/logs", methods=["GET"])
def get_attendance_logs():
    """Get stored attendance logs with pagination and filtering

    Query parameters:
    - device_id: Filter by device ID
    - limit: Max records to return (default: 100, max: 1000)
    - offset: Number of records to skip (default: 0)
    - user_id: Filter by user ID
    - date: Single date filter (YYYY-MM-DD) - for backward compatibility
    - start_date: Start of date range (YYYY-MM-DD)
    - end_date: End of date range (YYYY-MM-DD)

    Note: If both 'date' and 'start_date/end_date' are provided, 'date' takes precedence
    """
    try:
        from datetime import datetime

        # Query parameters
        device_id = request.args.get("device_id")
        limit = int(request.args.get("limit", 100))
        offset = int(request.args.get("offset", 0))
        user_id = request.args.get("user_id")
        date_str = request.args.get(
            "date"
        )  # Format: YYYY-MM-DD (single date - backward compatibility)
        start_date_str = request.args.get("start_date")  # Format: YYYY-MM-DD
        end_date_str = request.args.get("end_date")  # Format: YYYY-MM-DD

        # Validate limit
        limit = min(limit, 1000)  # Max 1000 records per request

        # Parse date filters
        start_date = None
        end_date = None

        # Priority 1: Single date (for backward compatibility)
        if date_str:
            try:
                target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                start_date = datetime.combine(target_date, datetime.min.time())
                end_date = datetime.combine(target_date, datetime.max.time())
            except ValueError:
                return jsonify(
                    {"success": False, "error": "Invalid date format. Use YYYY-MM-DD"}
                ), 400
        # Priority 2: Date range
        else:
            if start_date_str:
                try:
                    start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
                    start_date = datetime.combine(
                        start_date.date(), datetime.min.time()
                    )
                except ValueError:
                    return jsonify(
                        {
                            "success": False,
                            "error": "Invalid start_date format. Use YYYY-MM-DD",
                        }
                    ), 400

            if end_date_str:
                try:
                    end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
                    end_date = datetime.combine(end_date.date(), datetime.max.time())
                except ValueError:
                    return jsonify(
                        {
                            "success": False,
                            "error": "Invalid end_date format. Use YYYY-MM-DD",
                        }
                    ), 400

        # Get logs
        if user_id:
            logs = attendance_repo.get_by_user(user_id, limit)
        else:
            logs = attendance_repo.get_all(
                device_id, limit, offset, start_date, end_date
            )

        data = _map_user_details_to_logs(logs)

        # Get total count for proper pagination
        total_count = attendance_repo.get_total_count(device_id, start_date, end_date)

        return jsonify(
            {
                "success": True,
                "data": data,
                "pagination": {
                    "limit": limit,
                    "offset": offset,
                    "count": len(logs),
                    "total_count": total_count,
                },
                "date": date_str,
                "start_date": start_date_str,
                "end_date": end_date_str,
            }
        )

    except Exception as e:
        app_logger.error(f"Error getting attendance logs: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/attendance/unsynced", methods=["GET"])
def get_unsynced_logs():
    """Get attendance logs that haven't been synced to external system"""
    try:
        device_id = request.args.get("device_id")
        limit = int(request.args.get("limit", 1000))

        # Validate limit
        limit = min(limit, 1000)  # Max 1000 records per request

        logs = attendance_repo.get_unsynced_logs(device_id, limit)

        return jsonify(
            {
                "success": True,
                "data": [log.to_dict() for log in logs],
                "count": len(logs),
            }
        )

    except Exception as e:
        app_logger.error(f"Error getting unsynced logs: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/attendance/<int:log_id>/sync", methods=["PUT"])
def mark_log_as_synced(log_id: int):
    """Mark a specific attendance log as synced"""
    try:
        success = attendance_repo.mark_as_synced(log_id)

        if success:
            app_logger.info(f"Marked attendance log {log_id} as synced")
            return jsonify(
                {
                    "success": True,
                    "message": f"Attendance log {log_id} marked as synced",
                    "synced_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
            )
        else:
            return jsonify(
                {
                    "success": False,
                    "error": f"Attendance log {log_id} not found or already synced",
                }
            ), 404

    except Exception as e:
        app_logger.error(f"Error marking log {log_id} as synced: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/attendance/sync/batch", methods=["PUT"])
def mark_multiple_as_synced():
    """Mark multiple attendance logs as synced"""
    try:
        data = request.get_json()
        if not data or "log_ids" not in data:
            return jsonify(
                {"success": False, "error": "log_ids array is required in request body"}
            ), 400

        log_ids = data["log_ids"]
        if not isinstance(log_ids, list):
            return jsonify({"success": False, "error": "log_ids must be an array"}), 400

        synced_count = 0
        failed_ids = []

        for log_id in log_ids:
            try:
                if attendance_repo.mark_as_synced(int(log_id)):
                    synced_count += 1
                else:
                    failed_ids.append(log_id)
            except Exception as e:
                app_logger.error(f"Error syncing log {log_id}: {e}")
                failed_ids.append(log_id)

        app_logger.info(
            f"Batch sync completed: {synced_count} synced, {len(failed_ids)} failed"
        )

        return jsonify(
            {
                "success": True,
                "synced_count": synced_count,
                "failed_count": len(failed_ids),
                "failed_ids": failed_ids,
                "synced_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
        )

    except Exception as e:
        app_logger.error(f"Error in batch sync: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/attendance/stats", methods=["GET"])
def get_sync_stats():
    """Get sync statistics for attendance logs"""
    try:
        device_id = request.args.get("device_id")
        stats = attendance_repo.get_sync_stats(device_id)

        return jsonify({"success": True, "stats": stats, "device_id": device_id})

    except Exception as e:
        app_logger.error(f"Error getting sync stats: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# Daily Attendance Sync APIs


@bp.route("/attendance/sync-daily", methods=["POST"])
def sync_daily_attendance():
    """Sync daily attendance data with first checkin/last checkout logic"""
    try:
        data = request.get_json() or {}
        target_date = data.get("date")  # Optional: YYYY-MM-DD format
        device_id = data.get("device_id")  # Optional: specific device

        app_logger.info(
            f"Manual daily attendance sync triggered for date: {target_date or 'today'}, device: {device_id or 'all'}"
        )

        result = attendance_sync_service.sync_attendance_daily(
            target_date, device_id, ignore_error_limit=True
        )

        if result["success"]:
            dates_processed = result.get("dates_processed", [])
            date_info = (
                f"{len(dates_processed)} dates"
                if len(dates_processed) > 1
                else (dates_processed[0] if dates_processed else "today")
            )
            return jsonify(
                {
                    "success": True,
                    "message": f"Daily attendance sync completed for {date_info}",
                    "data": result,
                }
            )
        else:
            return jsonify(
                {
                    "success": False,
                    "error": result.get("error"),
                    "dates_processed": result.get("dates_processed", []),
                }
            ), 500

    except Exception as e:
        app_logger.error(f"Error in sync_daily_attendance: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/attendance/daily-preview", methods=["GET"])
def preview_daily_attendance():
    """Preview daily attendance data without sending to external API"""
    try:
        target_date = request.args.get("date")  # Optional: YYYY-MM-DD format
        device_id = request.args.get("device_id")  # Optional: specific device

        result = attendance_sync_service.get_daily_attendance_preview(
            target_date, device_id
        )

        return jsonify(result)

    except Exception as e:
        app_logger.error(f"Error in preview_daily_attendance: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# Scheduler Management APIs


@bp.route("/attendance/scheduler/status", methods=["GET"])
def get_scheduler_status():
    """Get scheduler status and job information"""
    try:
        status = scheduler_service.get_all_jobs()
        return jsonify({"success": True, "scheduler": status})

    except Exception as e:
        app_logger.error(f"Error getting scheduler status: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/attendance/scheduler/trigger", methods=["POST"])
def trigger_scheduler_job():
    """Manually trigger the daily attendance sync job"""
    try:
        data = request.get_json() or {}
        job_id = data.get("job_id", "daily_attendance_sync")

        app_logger.info(f"Manual trigger requested for job: {job_id}")

        result = scheduler_service.trigger_job_manually(job_id)

        if result["success"]:
            return jsonify(
                {
                    "success": True,
                    "message": result.get("message"),
                    "data": result.get("result"),
                }
            )
        else:
            return jsonify({"success": False, "error": result.get("error")}), 500

    except Exception as e:
        app_logger.error(f"Error triggering scheduler job: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/attendance/history", methods=["GET"])
def get_attendance_history():
    """Get attendance history with smart filtering (first checkin/last checkout or synced records)

    Logic:
    - Group by user + date
    - For each group:
      - If checkin has synced record, use it; else use first checkin
      - If checkout has synced record, use it; else use last checkout
    """
    try:
        from datetime import datetime

        # Query parameters
        device_id = request.args.get("device_id")
        date_str = request.args.get("date")  # Format: YYYY-MM-DD

        # Parse date or default to today
        if date_str:
            try:
                target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                return jsonify(
                    {"success": False, "error": "Invalid date format. Use YYYY-MM-DD"}
                ), 400
        else:
            target_date = datetime.now().date()

        # Get filtered logs from repository
        filtered_logs = attendance_repo.get_smart_filtered_by_date(
            target_date, device_id
        )

        # Map user names and avatars
        data = _map_user_details_to_logs(filtered_logs)

        return jsonify(
            {
                "success": True,
                "data": data,
                "date": target_date.strftime("%Y-%m-%d"),
                "count": len(data),
                "device_id": device_id,
            }
        )

    except Exception as e:
        app_logger.error(f"Error getting attendance history: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# Cleanup Management APIs


@bp.route("/attendance/cleanup/preview", methods=["GET"])
def preview_cleanup():
    """Preview what would be deleted in cleanup without actually deleting"""
    try:
        # Get retention days from query parameter (default: 365 = 1 year)
        retention_days = int(request.args.get("retention_days", 365))

        # Validate retention period
        if retention_days < 30:
            return jsonify(
                {
                    "success": False,
                    "error": "Retention period must be at least 30 days for safety",
                }
            ), 400

        result = attendance_cleanup_service.get_cleanup_preview(retention_days)

        return jsonify(result)

    except ValueError:
        return jsonify(
            {
                "success": False,
                "error": "Invalid retention_days parameter. Must be an integer.",
            }
        ), 400
    except Exception as e:
        app_logger.error(f"Error previewing cleanup: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/attendance/cleanup", methods=["POST"])
def cleanup_old_attendance():
    """Delete old synced and skipped attendance records

    Only deletes records with sync_status = 'synced' or 'skipped'.
    NEVER deletes 'pending' records to prevent data loss.
    """
    try:
        data = request.get_json() or {}

        # Get retention days from request body (default: 365 = 1 year)
        retention_days = data.get("retention_days", 365)

        # Validate retention period
        if retention_days < 30:
            return jsonify(
                {
                    "success": False,
                    "error": "Retention period must be at least 30 days for safety",
                }
            ), 400

        # Require explicit confirmation for safety
        confirmed = data.get("confirmed", False)
        if not confirmed:
            return jsonify(
                {
                    "success": False,
                    "error": 'Cleanup requires explicit confirmation. Set "confirmed": true in request body.',
                }
            ), 400

        app_logger.info(f"Manual cleanup triggered: retention_days={retention_days}")

        result = attendance_cleanup_service.cleanup_old_attendance(retention_days)

        if result["success"]:
            return jsonify(
                {"success": True, "message": result["message"], "data": result}
            )
        else:
            return jsonify(
                {
                    "success": False,
                    "error": result.get("error"),
                    "message": result.get("message"),
                }
            ), 500

    except Exception as e:
        app_logger.error(f"Error in cleanup_old_attendance: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/attendance/export-excel", methods=["GET"])
def export_attendance_excel():
    """Export attendance logs to Excel file with date range filter"""
    try:
        # Query parameters
        device_id = request.args.get("device_id")
        start_date_str = request.args.get("start_date")  # Format: YYYY-MM-DD
        end_date_str = request.args.get("end_date")  # Format: YYYY-MM-DD

        # Parse dates
        start_date = None
        end_date = None

        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
                start_date = datetime.combine(start_date.date(), datetime.min.time())
            except ValueError:
                return jsonify(
                    {
                        "success": False,
                        "error": "Invalid start_date format. Use YYYY-MM-DD",
                    }
                ), 400

        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
                end_date = datetime.combine(end_date.date(), datetime.max.time())
            except ValueError:
                return jsonify(
                    {
                        "success": False,
                        "error": "Invalid end_date format. Use YYYY-MM-DD",
                    }
                ), 400

        # Get all attendance logs (no limit for export)
        logs = attendance_repo.get_all(
            device_id=device_id,
            limit=100000,  # Large limit for export
            offset=0,
            start_date=start_date,
            end_date=end_date,
        )

        # Map user details
        enriched_logs = _map_user_details_to_logs(logs)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Dữ liệu chấm công"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        border_style = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        headers = [
            "STT",
            "Mã NV",
            "Tên nhân viên",
            "Họ tên đầy đủ",
            "Phòng ban",
            "Chức danh",
            "Giới tính",
            "Ngày vào làm",
            "Thời gian chấm công",
            "Phương thức",
            "Hành động",
            "Trạng thái đồng bộ",
            "Thiết bị",
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style

        # Define column widths
        column_widths = {
            "A": 6,  # STT
            "B": 12,  # Mã NV
            "C": 25,  # Tên NV
            "D": 30,  # Họ tên đầy đủ
            "E": 20,  # Phòng ban
            "F": 20,  # Chức danh
            "G": 12,  # Giới tính
            "H": 15,  # Ngày vào làm
            "I": 20,  # Thời gian
            "J": 15,  # Phương thức
            "K": 15,  # Hành động
            "L": 18,  # Trạng thái
            "M": 15,  # Thiết bị
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Method and action maps
        ATTENDANCE_METHOD_MAP = {
            0: "Mật khẩu",
            1: "Vân tay",
            2: "Thẻ",
            15: "Khuôn mặt",
        }

        PUNCH_ACTION_MAP = {
            0: "Vào ca",
            1: "Ra ca",
            2: "Bắt đầu nghỉ",
            3: "Kết thúc nghỉ",
            4: "Bắt đầu tăng ca",
            5: "Kết thúc tăng ca",
        }

        SYNC_STATUS_MAP = {
            "synced": "Đã đồng bộ",
            "pending": "Đang chờ",
            "skipped": "Đã bỏ qua",
            "error": "Lỗi",
        }

        def format_gender(value):
            if not value:
                return "-"
            normalized = str(value).strip().lower()
            if normalized in ["male", "nam", "m"]:
                return "Nam"
            if normalized in ["female", "nu", "nữ", "f"]:
                return "Nữ"
            if normalized in ["other", "khac", "khác", "unspecified", "unknown"]:
                return "Khác"
            return value

        def format_date(date_str):
            if not date_str:
                return "-"
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
                return dt.strftime("%d/%m/%Y")
            except:
                return date_str

        def format_timestamp(timestamp_str):
            """Parse timestamp with flexible format (with or without milliseconds)"""
            if not timestamp_str:
                return "-"
            try:
                # Try with milliseconds first
                dt = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S.%f")
                return dt.strftime("%d/%m/%Y %H:%M:%S")
            except ValueError:
                try:
                    # Try without milliseconds
                    dt = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
                    return dt.strftime("%d/%m/%Y %H:%M:%S")
                except ValueError:
                    # Return as-is if can't parse
                    return timestamp_str

        # Write data rows
        for idx, log in enumerate(enriched_logs, 1):
            row_num = idx + 1

            # Get sync status
            sync_status = log.get("sync_status") or (
                "synced" if log.get("is_synced") else "pending"
            )

            row_data = [
                idx,  # STT
                log.get("employee_code") or "-",
                log.get("name") or "Unknown User",
                log.get("full_name") or "-",
                log.get("department") or "-",
                log.get("position") or "-",
                format_gender(log.get("gender")),
                format_date(log.get("hire_date")),
                format_timestamp(log.get("timestamp")),
                ATTENDANCE_METHOD_MAP.get(log["method"], "Không xác định"),
                PUNCH_ACTION_MAP.get(log["action"], "Không xác định"),
                SYNC_STATUS_MAP.get(sync_status, "Không xác định"),
                log.get("device_id") or "-",
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border_style
                cell.alignment = Alignment(vertical="center")

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate filename
        date_range = ""
        if start_date_str and end_date_str:
            date_range = f"_{start_date_str}_den_{end_date_str}"
        elif start_date_str:
            date_range = f"_tu_{start_date_str}"
        elif end_date_str:
            date_range = f"_den_{end_date_str}"

        filename = (
            f"cham_cong{date_range}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        app_logger.info(f"Exported {len(enriched_logs)} attendance records to Excel")

        return send_file(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        app_logger.error(f"Error exporting attendance to Excel: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
